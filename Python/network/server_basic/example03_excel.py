import http.server
import socketserver
import io

import xlsxwriter

# openpyxl로 하기위한 모듈
from openpyxl import Workbook
from tempfile import NamedTemporaryFile

PORT = 8001

class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        # flush 방식
        # 다음에 출력할 문자들을 잡아준다
        # 잡아준 문자들을 보내기만 하면 된다
        # 화면에 출력되는 것들을 잡아서 변수에 저장한다.
        # 메모리상에 저장한다

        # output = io.BytesIO()
        # # xlsxwriter로 만든 데이터를 BytesIO라는 데이터 스트림으로 변환해서 출력한 것이다
        # workbook = xlsxwriter.Workbook(output, {'in_memory' : True})
        # worksheet = workbook.add_worksheet()
        # worksheet.write(0,0, "EXCEL TEST")
        # workbook.close()
        # output.seek(0)

        # openpyxl 로 할수있지 않을까?
        with NamedTemporaryFile() as tmp:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "data example"
            wb.save(tmp.name)

            # 파일 포인터를 처음으로 옮기고 읽기 위해 seek(0)을 한다
            tmp.seek(0)
            stream = tmp.read()

        self.send_response(200)
        self.send_header('Content-Disposition', 'attachment; filename=test.xlsx')
        self.send_header('Content-type',
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.end_headers()
        self.wfile.write(stream)

        return

print("serving at port", PORT)
httpd = socketserver.TCPServer(('', PORT), Handler)
httpd.serve_forever()