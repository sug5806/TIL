# 파일 형식중 하나이다 : html, xml, json
import json

import requests
from openpyxl import Workbook

# AJAX로 통신을 하면 JSON방식으로 한다.

# 헤더를 추가하지않으면 402(forbidden)이 뜬다
custom_headers = {
    'referer': 'http://daum.net/',
    'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Whale/1.5.71.11 Safari/537.36',
}

count = 2


def code():
    url = "http://finance.daum.net/api/quote/A048410/days?symbolCode=A048410&page=1&perPage=100"

    req = requests.get(url, headers=custom_headers)

    if req.status_code == requests.codes.ok:
        print("접속 성공")
        stock_data = json.loads(req.text)  # stock_data는 dictionary 형태이다
        for daily_data in stock_data['data'][:5]:
            print(daily_data['date'], daily_data['tradePrice'])


# 크톤탭에 작업 등록하는 방법 블로그에 정리하기
# 작업 주기
# 실행할 작업
# 옵션

"""
메뉴 1: 급등 항목 10개의 현재가 출력
"""

def real_boost():
    url = 'http://finance.daum.net/api/search/ranks?limit=10'

    # req.text를 하면 JSON 형태의 출력물이 나온다
    req = requests.get(url, headers=custom_headers)
    if req.status_code == requests.codes.ok:
        print("접속 성공")
        # json.loads를 통해 key : value 형태로 가져온다
        stock_data = json.loads(req.text)
        for daily_data in stock_data['data']:
            print(daily_data['rank'], daily_data['symbolCode'], daily_data['code'], daily_data['name'],
                  daily_data['tradePrice'])
        print()


# 메뉴 2 : 종목코드 입력받고 해당 종목의 1페이지 주가 출력
# 메뉴 2-2 : 데이터 엑셀로 저장하기
def first_page(code):
    url = 'http://finance.daum.net/api/quotes/' + code + '?summary=false&changeStatistics=true'
    # 내가 주소로 get 요청을 보냈고 그에 대한 응답으로 200이라는 상태코드와 같이 왔다
    req = requests.get(url, headers=custom_headers)

    if req.status_code == requests.codes.ok:
        # print("접속완료")
        stock_data = json.loads(req.text)
        print(f"종목 명: {stock_data['name']}, 종목 코드 : {stock_data['symbolCode']}, 현재 가격: {stock_data['openingPrice']},"
              f" 전일가 : {stock_data['prevClosingPrice']}, 고가 : {stock_data['highPrice']}, 저가 : {stock_data['lowPrice']}")
        print()

        yn = input("파일을 저장하시겠습니까? y/n")
        if yn == "y":
            global count
            # 엑셀파일 할당
            wb = Workbook()
            # default 시트 할당
            ws = wb.active
            # sheet의 이름은 Stock
            ws.title = "Stock"

            ws.cell(row=1, column=1, value="종목명")
            ws.cell(row=1, column=2, value="종목 코드")
            ws.cell(row=1, column=3, value="현재 가격")
            ws.cell(row=1, column=4, value="전일가")
            ws.cell(row=1, column=5, value="고가")
            ws.cell(row=1, column=6, value="저가")

            ws.cell(row=count, column=1, value=stock_data['name'])
            ws.cell(row=count, column=2, value=stock_data['symbolCode'])
            ws.cell(row=count, column=3, value=stock_data['openingPrice'])
            ws.cell(row=count, column=4, value=stock_data['prevClosingPrice'])
            ws.cell(row=count, column=5, value=stock_data['highPrice'])
            ws.cell(row=count, column=6, value=stock_data['lowPrice'])
            count += 1

            # stock.xlsx로 저장
            wb.save("stock.xlsx")
            wb.close()

if __name__ == "__main__":
    while True:
        print("메뉴를 선택해주세요 : ")
        print("1 : 급등항목 10개의 현재가 조회")
        print("0 : 종료")
        print("2 : 종목코드 입력하고 해당 종목의 주가 출력")
        choice = int(input());

        if not choice:
            break
        elif choice == 1:
            real_boost()
            print()
        elif choice == 2:
            code = input("종목 코드를 입력해주세요 : ")
            print()
            first_page(code)

"""
구현하는 순서
1. 메뉴 구현하기
2. 1번 메뉴에대한 크롤러
2-1. 조회 급등 항목 10개 찾기
2-2. 그 항목 1개에 대한 데이터 찾아서 출력
2-3. 총 10개 항목에 대한 데이터 찾아서 출력

3. 2번 메뉴에 대한 크롤러
3-1. 아무 항목이나 - 페이지 주가 출력
3-2. 사용자에게 종목 코드 입력받아서 출력하기"""