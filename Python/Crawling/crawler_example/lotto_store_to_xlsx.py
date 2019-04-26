# 1. 로또 번호 크롤러
#  * 1. 1회차분 Crawling
#  * 2. 다회차 Crawling


# 2. 엑셀에 데이터 저장
#    * 1. 아무값이나 엑셀에 써보기
#    * 2. 특정 회차 번호 저장해보기
#    * 3. 다회차 저장해보기


# 분기를 할때는 프론트(주소)에서 분류 되는것과 백엔드(뒤)에서 분류되는것으로 2가지를 나뉘어져 있다
# 1~855회차 정보 수집
# 엑셀 파일 저장하기
# 1 : ?, 2 : ?, 3 : ?, 4 : ? 번호가 지금까지 몇번이 나왔나

import requests

from bs4 import BeautifulSoup

# 파이썬으로 txt, csv, excel 다루기

from openpyxl import Workbook

# Workbook : 액셀 파일
# WorkSheet : 액셀의 시트

# 엑셀 파일
wb = Workbook()
# 현재 선택되어있는 엑셀시트 선택
ws1 = wb.active
# ws1.cell(row=1, column=1, value="test") # A1
# ws1['A1'] = "aaa" 이렇게도 가능

# 엑셀시트 만들기
ws2 = wb.create_sheet(title="second sheet")
# ws2.cell(row=3, column=7, value="두번째 시트")

# comprehension
count = {x: 0 for x in range(1, 46)}

for num in range(1, 31):
    url3 = "https://search.daum.net/search?w=tot&rtmaxcoll=LOT&DA=LOT&q=" + str(num) + "회차%20로또"
    req3 = requests.get(url3)

    if req3.status_code == requests.codes.ok:
        # print("접속 성공")
        html = BeautifulSoup(req3.text, "html.parser")

        number = html.select('span.img_lotto')
        ws2.cell(row=num, column=1, value="{}회차".format(num))
        col = 2
        for num2 in number[:7]:
            print(num2.text, end=", ")
            # count.update({int(num2.text) : int(count.get(num2.text, 0))+1})
            count[int(num2.text)] += 1
            ws2.cell(row=num, column=col + 1, value=num2.text)
            col += 1
        ws2.cell(row=num, column=col + 1, value="보너스 : {}".format(num2.text))
    print()

print(count)
for i in range(1, 46):
    ws1.cell(row=i, column=1, value=i)
    ws1.cell(row=i, column=2, value=f'{count[i]} 번')

wb.save("test.xlsx")
wb.close()